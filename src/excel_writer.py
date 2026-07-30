"""Escritura del Excel de rendición siguiendo el formato exacto de la plantilla.

Parte de una copia de la plantilla (preserva hojas, formato, encabezados, dropdowns
de Currency/Expense Type y la fórmula de la fila de total), limpia las filas de datos
y escribe una fila por boleta procesada con éxito.

- Currency: solo códigos de la lista fija de la plantilla (ver currency.py).
- FX: siempre 1 (v1 no hace conversión de divisas; el usuario ajusta el FX real a mano).
- Amount in CLP: fórmula de Excel "=FX*Amount" (se recalcula sola si el usuario edita FX).
- Comments: en blanco (lo completa el usuario).
- Date Submitted (D6) y Date del bloque "Approved by:" (F39): fecha de ejecución del
  programa, no la fecha de cada boleta (esa es la columna Date por fila, COL_DATE).
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from xml.etree import ElementTree as ET
from datetime import date as date_cls
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

# Columnas de la hoja "Expense Report" (fila 8 = encabezado).
COL_ITEM = 2       # B
COL_DATE = 3       # C
COL_AMOUNT = 4     # D
COL_CURRENCY = 5   # E
COL_FX = 6         # F
COL_AMOUNT_CLP = 7  # G
COL_EXPENSE_TYPE = 8  # H
COL_COMMENTS = 9   # I

DEFAULT_FX = 1.0

# Campos de fecha del encabezado/metadata del formulario (no confundir con COL_DATE,
# la columna de fecha por boleta). Se estampan con la fecha de ejecución cada corrida.
DATE_SUBMITTED_CELL = "D6"   # etiqueta "Date Submitted:" en B6
APPROVAL_DATE_CELL = "F39"   # etiqueta "Date   :" en E39, dentro del bloque "Approved by:"

# La plantilla trae los dropdowns de Currency/Expense Type acotados a las filas 9-32
# (validación de datos 'Cheat Sheet'!$A$37:$A$43 y $A$2:$A$23) y una fila de total en
# la 33 (`=SUM(G9:G32)`). No hay que limpiar ni escribir más allá de esa fila.
DEFAULT_LAST_DATA_ROW = 32

_EXT_LST_RE = re.compile(r"<extLst>.*?</extLst>", re.S)


class TemplateCapacityError(ValueError):
    """El número de boletas OK excede las filas de datos disponibles en la plantilla."""


@dataclass
class ExpenseRow:
    date: Optional[date_cls]
    amount: float
    currency: str
    expense_type: Optional[str]
    source_file: str
    fx: float = DEFAULT_FX
    comments: Optional[str] = None


def _set_cell(ws, row: int, col: int, value) -> None:
    # ws.cell(row, column, value=None) es un no-op en openpyxl cuando value es None
    # (lo interpreta como "no cambiar"), así que hay que asignar .value directamente
    # para poder limpiar celdas.
    ws.cell(row=row, column=col).value = value


def _clear_data_rows(ws, first_row: int, last_row: int) -> None:
    for row in range(first_row, last_row + 1):
        for col in range(COL_ITEM, COL_COMMENTS + 1):
            _set_cell(ws, row, col, None)


_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def _sheet_xml_path(zf: zipfile.ZipFile, sheet_name: str) -> str:
    """Resuelve el nombre de una hoja (ej. 'Expense Report') a su archivo sheetN.xml,
    parseando workbook.xml + workbook.xml.rels. Independiente del orden de atributos
    y del formato de Target (relativo o con "/" inicial), que difieren entre la
    plantilla original y lo que reescribe openpyxl al guardar.
    """
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rid = None
    for sheet_el in workbook_root.iter(f"{{{_NS_MAIN}}}sheet"):
        if sheet_el.get("name") == sheet_name:
            rid = sheet_el.get(f"{{{_NS_R}}}id")
            break
    if not rid:
        raise ValueError(f"No se encontró la hoja '{sheet_name}' en workbook.xml")

    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target = None
    for rel_el in rels_root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        if rel_el.get("Id") == rid:
            target = rel_el.get("Target")
            break
    if not target:
        raise ValueError(f"No se encontró la relación '{rid}' en workbook.xml.rels")

    target = target.lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _extract_ext_lst(template_path: Path, sheet_name: str) -> Optional[str]:
    """Lee el bloque <extLst> (dropdowns x14:dataValidation) de la hoja en la plantilla original."""
    with zipfile.ZipFile(template_path) as zf:
        sheet_path = _sheet_xml_path(zf, sheet_name)
        sheet_xml = zf.read(sheet_path).decode("utf-8")
    match = _EXT_LST_RE.search(sheet_xml)
    if not match:
        return None
    # xr:uid depende de xmlns:xr declarado en la raíz <worksheet> de la plantilla
    # original, que openpyxl no reescribe. Sin esa declaración el prefijo queda sin
    # vincular y el XML resultante es inválido. Los uid son solo IDs de tracking de
    # revisión de Excel, no afectan el funcionamiento del dropdown: se quitan.
    return re.sub(r'\s+xr:uid="[^"]*"', "", match.group(0))


def _restore_ext_lst(output_path: Path, sheet_name: str, ext_lst_xml: str) -> None:
    """Reinyecta el <extLst> que openpyxl descarta al guardar (no soporta esa extensión),
    para que los dropdowns de Currency/Expense Type sigan funcionando en el archivo final.
    """
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    with zipfile.ZipFile(output_path) as zin:
        sheet_path = _sheet_xml_path(zin, sheet_name)
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == sheet_path:
                    text = data.decode("utf-8")
                    if "<extLst>" not in text:
                        text = text.replace("</worksheet>", f"{ext_lst_xml}</worksheet>")
                        data = text.encode("utf-8")
                zout.writestr(item, data)
    tmp_path.replace(output_path)


def write_expense_report(
    rows: List[ExpenseRow], config: dict, output_path: Path
) -> Path:
    excel_cfg = config.get("excel", {})
    template_path = Path(excel_cfg.get("template_path"))
    sheet_name = excel_cfg.get("sheet_name", "Expense Report")
    first_row = excel_cfg.get("first_data_row", 9)
    last_data_row = excel_cfg.get("last_data_row", DEFAULT_LAST_DATA_ROW)

    capacity = last_data_row - first_row + 1
    if len(rows) > capacity:
        raise TemplateCapacityError(
            f"{len(rows)} boletas OK exceden las {capacity} filas de datos disponibles "
            f"en la plantilla (filas {first_row}-{last_data_row})."
        )

    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name]

    _clear_data_rows(ws, first_row, last_data_row)

    if excel_cfg.get("employee_name"):
        ws["D4"] = excel_cfg["employee_name"]
    if excel_cfg.get("cost_centre"):
        ws["D5"] = excel_cfg["cost_centre"]

    # Fecha de ejecución en los dos campos de fecha del encabezado ("Date Submitted"
    # y "Date" del bloque de firma). Se respeta el number_format que ya trae cada
    # celda en la plantilla (no se toca): D6 se ve estilo "27-May-26", F39 "mm-dd-yy".
    execution_date = date_cls.today()
    ws[DATE_SUBMITTED_CELL] = execution_date
    ws[APPROVAL_DATE_CELL] = execution_date

    amount_col_letter = get_column_letter(COL_AMOUNT)
    fx_col_letter = get_column_letter(COL_FX)

    for i, row in enumerate(rows):
        r = first_row + i
        _set_cell(ws, r, COL_ITEM, i + 1)
        _set_cell(ws, r, COL_DATE, row.date)
        _set_cell(ws, r, COL_AMOUNT, float(row.amount))
        _set_cell(ws, r, COL_CURRENCY, row.currency)
        _set_cell(ws, r, COL_FX, row.fx)
        _set_cell(
            ws, r, COL_AMOUNT_CLP, f"={fx_col_letter}{r}*{amount_col_letter}{r}"
        )
        _set_cell(ws, r, COL_EXPENSE_TYPE, row.expense_type)
        _set_cell(ws, r, COL_COMMENTS, row.comments)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    ext_lst = _extract_ext_lst(template_path, sheet_name)
    if ext_lst:
        _restore_ext_lst(output_path, sheet_name, ext_lst)

    return output_path
