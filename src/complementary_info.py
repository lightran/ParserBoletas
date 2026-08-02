"""Calcula el FX real de cada moneda extranjera (regla de 3, dos pasos) a partir del
monto en USD que el banco cobró en la tarjeta, y lo documenta en la pestaña
"Complementary info" del Excel de rendición — una pestaña clonada por moneda
extranjera que lo necesite (ver `CurrencyConversion`), o ninguna si no aplica.

Regla (ejemplo real, boleta en soles peruanos):
    Dato A = monto de la boleta en moneda origen (223.50 PEN)
    Dato B = USD que el banco cobró por ese movimiento en la tarjeta (67.41 USD)
    Dato C = tipo de cambio USD -> CLP del banco (922)
    Paso 1: FX(origen->USD) = Dato B / Dato A          (0.301611)
    Paso 2: FX(origen->CLP) = FX(origen->USD) * Dato C (278.09 ≈ 278)
FX(origen->CLP) (Paso 2) es el valor que se carga en la columna FX del reporte para
todas las filas de esa moneda.

La pestaña "Complementary info" de la plantilla ya trae, como ejemplo de referencia,
las dos tablas de este cálculo (celdas P23:R31) y 3 imágenes: la boleta+voucher (zona
derecha, T9:AB52 — la única que esta app reemplaza, por la boleta seleccionada), y dos
screenshots del banco (listado de movimientos y tipo de cambio, zonas izquierda/
superior — la app nunca las toca, las actualiza el usuario a mano) conectadas a las
tablas con flechas azules.

Dos limitaciones de openpyxl obligan a un manejo especial (mismo espíritu que el
<extLst> de los dropdowns en excel_writer.py):
- `Workbook.copy_worksheet()` no copia imágenes/dibujos — se re-embeben a mano con
  las mismas coordenadas de anclaje que trae la plantilla (`_reattach_images`).
- Al guardar CUALQUIER hoja con dibujos, openpyxl descarta los conectores de flecha
  (`<xdr:cxnSp>`) porque no los modela — se reinyectan como XML crudo después de
  guardar (`reinject_arrows`), igual que el `<extLst>`.
"""

from __future__ import annotations

import posixpath
import re
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import List, Optional
from xml.etree import ElementTree as ET

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

from audit_writer import MAX_IMAGE_WIDTH_PX, load_original_image_for_embedding

SOURCE_SHEET_NAME = "Complementary info"

# Celdas de las dos tablas de conversión en la plantilla (ver docstring del módulo).
CELL_DATO_A = "P24"          # monto de la boleta en moneda origen
CELL_DATO_B = "Q24"          # USD cobrado por el banco
CELL_STEP1_RESULT = "R24"    # fórmula "=Q24/P24" ya en la plantilla, no se toca
CELL_DATO_C = "Q30"          # TC USD -> CLP del banco
CELL_STEP1_COPY = "P31"      # referencia al resultado del Paso 1 ("=R24")
CELL_STEP2_RESULT = "R30"    # fórmula "=Q30*P31" ya en la plantilla, no se toca

# Columna (0-indexed, como usa openpyxl para anclas de dibujo) donde arranca la
# imagen de la boleta+voucher en la plantilla (zona derecha, T9:AB52). Las otras
# imágenes (screenshots del banco) se reinsertan sin cambios.
_BOLETA_IMAGE_FROM_COL = 19

_MAX_SHEET_NAME_LEN = 31

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"


def compute_real_fx(dato_a: float, dato_b: float, dato_c: float) -> tuple[float, float]:
    """Regla de 3 en dos pasos. Devuelve (FX origen->USD, FX origen->CLP)."""
    fx_step1 = dato_b / dato_a
    fx_step2 = fx_step1 * dato_c
    return fx_step1, fx_step2


@dataclass
class CurrencyConversion:
    currency: str
    file_path: Path
    dato_a: float
    dato_b: float
    dato_c: float

    @property
    def fx_step1(self) -> float:
        return compute_real_fx(self.dato_a, self.dato_b, self.dato_c)[0]

    @property
    def fx_step2(self) -> float:
        return compute_real_fx(self.dato_a, self.dato_b, self.dato_c)[1]

    def sheet_title(self) -> str:
        title = f"{SOURCE_SHEET_NAME} - {self.currency}"
        return title[:_MAX_SHEET_NAME_LEN]


def _sheet_xml_path(zf: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rid = None
    for sheet_el in workbook_root.iter(f"{{{_NS_MAIN}}}sheet"):
        if sheet_el.get("name") == sheet_name:
            rid = sheet_el.get(f"{{{_NS_R}}}id")
            break
    if not rid:
        raise ValueError(f"No se encontró la hoja '{sheet_name}' en workbook.xml")

    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target = None
    for rel_el in rels_root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        if rel_el.get("Id") == rid:
            target = rel_el.get("Target")
            break
    if not target:
        raise ValueError(f"No se encontró la relación '{rid}' en workbook.xml.rels")

    target = target.lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _sheet_drawing_path(zf: zipfile.ZipFile, sheet_xml_path: str) -> Optional[str]:
    # sheet_xml_path ej. "xl/worksheets/sheet2.xml" -> rels en
    # "xl/worksheets/_rels/sheet2.xml.rels".
    sheet_dir, sheet_file = posixpath.split(sheet_xml_path)
    rels_path = posixpath.join(sheet_dir, "_rels", sheet_file + ".rels")
    if rels_path not in zf.namelist():
        return None
    rels_root = ET.fromstring(zf.read(rels_path))
    for rel_el in rels_root.iter(f"{{{_NS_PKG_REL}}}Relationship"):
        if rel_el.get("Type", "").endswith("/drawing"):
            target = rel_el.get("Target")
            if target.startswith("/"):
                return target.lstrip("/")
            # Relativo al directorio de la hoja (ej. "../drawings/drawing2.xml"
            # relativo a "xl/worksheets/").
            return posixpath.normpath(posixpath.join(sheet_dir, target))
    return None


def _write_conversion_values(ws, conversion: CurrencyConversion, fx_number_format: Optional[str]) -> None:
    ws[CELL_DATO_A] = float(conversion.dato_a)
    ws[CELL_DATO_B] = float(conversion.dato_b)
    ws[CELL_DATO_C] = float(conversion.dato_c)
    ws[CELL_STEP1_COPY] = f"={CELL_STEP1_RESULT}"
    if fx_number_format:
        # La plantilla trae R30 con formato "0" (sin decimales) — inconsistente con
        # el formato de la columna FX en "Expense Report" (2 decimales), que es donde
        # termina usándose este mismo valor. Se iguala al formato real de esa columna
        # para que el mismo número se vea igual en las dos hojas.
        ws[CELL_STEP2_RESULT].number_format = fx_number_format


def _reattach_images(
    ws, template_images: list, conversion: CurrencyConversion, pdf_render_dpi: int
) -> None:
    for data, frm, to in template_images:
        if frm[0] == _BOLETA_IMAGE_FROM_COL:
            receipt = load_original_image_for_embedding(
                conversion.file_path, pdf_render_dpi, MAX_IMAGE_WIDTH_PX
            )
            if receipt is not None:
                buffer = BytesIO()
                receipt.save(buffer, format="PNG")
                buffer.seek(0)
                data = buffer.read()

        new_img = XLImage(BytesIO(data))
        anchor = TwoCellAnchor(
            editAs="oneCell",
            _from=AnchorMarker(*frm),
            to=AnchorMarker(*to),
        )
        ws.add_image(new_img, anchor)


def apply_conversions(
    wb,
    conversions: List[CurrencyConversion],
    pdf_render_dpi: int = 300,
    fx_number_format: Optional[str] = None,
) -> List[str]:
    """Muta `wb` (ya cargado desde la plantilla): reemplaza la pestaña
    "Complementary info" por una copia por cada conversión (ninguna si `conversions`
    está vacío). `fx_number_format` es el formato numérico de la columna FX en
    "Expense Report" — se le aplica al resultado final (Paso 2) para que el mismo
    valor se vea con la misma cantidad de decimales en las dos hojas. Devuelve los
    títulos de las pestañas creadas, para que `reinject_arrows` sepa cuáles parchar
    después de guardar."""
    template_ws = wb[SOURCE_SHEET_NAME]
    template_images = [
        (img._data(), (img.anchor._from.col, img.anchor._from.colOff, img.anchor._from.row, img.anchor._from.rowOff),
         (img.anchor.to.col, img.anchor.to.colOff, img.anchor.to.row, img.anchor.to.rowOff))
        for img in template_ws._images
    ]

    created_titles: List[str] = []
    for conversion in conversions:
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = conversion.sheet_title()
        _write_conversion_values(new_ws, conversion, fx_number_format)
        _reattach_images(new_ws, template_images, conversion, pdf_render_dpi)
        created_titles.append(new_ws.title)

    del wb[SOURCE_SHEET_NAME]

    if created_titles:
        # Reordena para que las pestañas nuevas queden justo después de "Expense
        # Report" (no al final, donde copy_worksheet las agrega por defecto).
        expense_report_idx = wb.sheetnames.index("Expense Report")
        created_sheets = [wb[title] for title in created_titles]
        for sheet in created_sheets:
            wb._sheets.remove(sheet)
        for offset, sheet in enumerate(created_sheets):
            wb._sheets.insert(expense_report_idx + 1 + offset, sheet)

    return created_titles


def _extract_arrow_blocks(template_path: Path) -> str:
    with zipfile.ZipFile(template_path) as zf:
        sheet_xml_path = _sheet_xml_path(zf, SOURCE_SHEET_NAME)
        drawing_path = _sheet_drawing_path(zf, sheet_xml_path)
        if not drawing_path:
            return ""
        drawing_xml = zf.read(drawing_path).decode("utf-8")

    blocks = re.findall(r"<xdr:twoCellAnchor>.*?</xdr:twoCellAnchor>", drawing_xml, re.S)
    arrow_blocks = [b for b in blocks if "cxnSp" in b]
    return "".join(
        b.replace("<xdr:twoCellAnchor>", f'<xdr:twoCellAnchor xmlns:xdr="{_XDR_NS}">', 1)
        for b in arrow_blocks
    )


def reinject_arrows(output_path: Path, template_path: Path, sheet_names: List[str]) -> None:
    """Reinyecta las flechas conectoras (`<xdr:cxnSp>`) que openpyxl descarta al
    guardar, en cada pestaña de `sheet_names` del archivo ya guardado en
    `output_path`. No-op si `sheet_names` está vacío."""
    if not sheet_names:
        return

    injected = _extract_arrow_blocks(template_path)
    if not injected:
        return

    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    with zipfile.ZipFile(output_path) as zin:
        drawing_paths = set()
        for name in sheet_names:
            sheet_xml_path = _sheet_xml_path(zin, name)
            drawing_path = _sheet_drawing_path(zin, sheet_xml_path)
            if drawing_path:
                drawing_paths.add(drawing_path)

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in drawing_paths:
                    text = data.decode("utf-8")
                    text = text.replace("</wsDr>", injected + "</wsDr>")
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    tmp_path.replace(output_path)
