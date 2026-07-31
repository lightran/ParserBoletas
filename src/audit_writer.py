"""Genera el reporte de auditoría como un Excel (.xlsx): una pestaña por boleta
marcada para revisión, con el motivo, los valores extraídos relevantes y la
imagen ORIGINAL de la boleta embebida (nunca la versión preprocesada/binarizada,
para que se lea de forma natural). Si no hubo boletas para revisión en la
corrida, no se genera el archivo.
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PIL import Image as PILImage

import preprocess
from extract import ExtractionResult
from validate import ValidationResult

INDEX_SHEET_NAME = "Índice"
MAX_SHEET_NAME_LEN = 31
MAX_IMAGE_WIDTH_PX = 700  # tamaño de visualización razonable; no infla el archivo

_FORBIDDEN_SHEET_CHARS_RE = re.compile(r'[:\\/?*\[\]]')

ReviewCase = Tuple[Path, ExtractionResult, ValidationResult]


def sanitize_sheet_name(raw_name: str, existing_names) -> str:
    """Deriva un nombre de pestaña válido y único a partir de `raw_name`.

    Excel limita los nombres a 31 caracteres y prohíbe : \\ / ? * [ ]. Si el
    nombre saneado/truncado ya existe, se le agrega un sufijo numérico.
    """
    name = _FORBIDDEN_SHEET_CHARS_RE.sub("_", raw_name).strip()
    if not name:
        name = "Revision"
    name = name[:MAX_SHEET_NAME_LEN]

    if name not in existing_names:
        return name

    suffix_n = 2
    while True:
        suffix = f"_{suffix_n}"
        candidate = name[: MAX_SHEET_NAME_LEN - len(suffix)] + suffix
        if candidate not in existing_names:
            return candidate
        suffix_n += 1


def _load_original_image_for_embedding(
    file_path: Path, pdf_render_dpi: int, max_width_px: int
) -> Optional[PILImage.Image]:
    """Carga la imagen ORIGINAL (sin preprocesar) de la boleta, rasterizando PDFs,
    y la reescala a `max_width_px` de ancho manteniendo proporción."""
    try:
        pages = preprocess.load_pages(file_path, pdf_render_dpi)
    except Exception:
        return None
    if not pages:
        return None

    img = pages[0].convert("RGB")
    w, h = img.size
    if w > max_width_px:
        scale = max_width_px / w
        img = img.resize((max_width_px, round(h * scale)), PILImage.LANCZOS)
    return img


def _add_review_sheet(wb, file_path: Path, result: ExtractionResult, validation: ValidationResult, config: dict) -> str:
    existing_names = {ws.title for ws in wb.worksheets}
    sheet_name = sanitize_sheet_name(file_path.stem, existing_names)
    ws = wb.create_sheet(sheet_name)

    fields = [
        ("Archivo de origen", file_path.name),
        ("Motivo de revisión", "; ".join(validation.reasons)),
        ("Monto", float(result.amount) if result.amount is not None else None),
        ("Moneda", result.currency),
        ("Fecha", result.date),
    ]
    for i, (label, value) in enumerate(fields, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    pdf_render_dpi = config.get("preprocess", {}).get("pdf_render_dpi", 300)
    image = _load_original_image_for_embedding(file_path, pdf_render_dpi, MAX_IMAGE_WIDTH_PX)
    image_row = len(fields) + 2
    if image is not None:
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        ws.add_image(XLImage(buffer), f"A{image_row}")
    else:
        ws.cell(row=image_row, column=1, value="(No se pudo cargar la imagen original de la boleta)")

    return sheet_name


def write_audit_report(
    review_cases: List[ReviewCase], config: dict, output_path: Path
) -> Optional[Path]:
    """Escribe el .xlsx de auditoría (una pestaña por caso + índice), o no genera
    nada si `review_cases` está vacío. En ese caso, si existe un archivo de una
    corrida anterior en `output_path`, se borra: su ausencia debe seguir siendo
    una señal confiable de "esta corrida no tuvo casos para revisión"."""
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        corrected = output_path.with_suffix(".xlsx")
        print(
            f"Aviso: el reporte de auditoría siempre se guarda en formato .xlsx — "
            f"'{output_path}' no termina en .xlsx, se usará '{corrected}' en su lugar."
        )
        output_path = corrected

    if not review_cases:
        if output_path.exists():
            output_path.unlink()
        return None

    wb = openpyxl.Workbook()
    index_ws = wb.active
    index_ws.title = INDEX_SHEET_NAME
    index_ws["A1"] = "Archivo"
    index_ws["B1"] = "Motivo de revisión"
    index_ws["A1"].font = Font(bold=True)
    index_ws["B1"].font = Font(bold=True)
    index_ws.column_dimensions["A"].width = 35
    index_ws.column_dimensions["B"].width = 60

    for i, (file_path, result, validation) in enumerate(review_cases):
        sheet_name = _add_review_sheet(wb, file_path, result, validation, config)

        row = i + 2
        file_cell = index_ws.cell(row=row, column=1, value=file_path.name)
        file_cell.hyperlink = f"#'{sheet_name}'!A1"
        file_cell.font = Font(color="0563C1", underline="single")
        index_ws.cell(row=row, column=2, value="; ".join(validation.reasons))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
