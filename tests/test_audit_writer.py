import zipfile
from pathlib import Path

import openpyxl
import pytest

from audit_writer import (
    INDEX_SHEET_NAME,
    MAX_SHEET_NAME_LEN,
    sanitize_sheet_name,
    write_audit_report,
)
from extract import ExtractionResult
from validate import ValidationResult

PROJECT_ROOT = Path(__file__).resolve().parent.parent
BOLETAS_DIR = PROJECT_ROOT / "boletas"

CONFIG = {"preprocess": {"pdf_render_dpi": 150}}


def _make_result(**overrides) -> ExtractionResult:
    defaults = dict(amount=235.40, currency="PEN", date="2026-06-10")
    defaults.update(overrides)
    return ExtractionResult(**defaults)


# --- sanitize_sheet_name ---


def test_sanitize_sheet_name_replaces_forbidden_characters():
    name = sanitize_sheet_name("Taxi: SCL/Home [Invoice]?*", existing_names=set())
    assert not any(ch in name for ch in ':\\/?*[]')


def test_sanitize_sheet_name_truncates_to_max_length():
    long_name = "Este es un nombre de archivo de boleta extremadamente largo y descriptivo"
    name = sanitize_sheet_name(long_name, existing_names=set())
    assert len(name) <= MAX_SHEET_NAME_LEN


def test_sanitize_sheet_name_dedupes_on_collision():
    existing = {"Taxi to Kyndryl"}
    name = sanitize_sheet_name("Taxi to Kyndryl", existing_names=existing)
    assert name != "Taxi to Kyndryl"
    assert name not in existing
    assert len(name) <= MAX_SHEET_NAME_LEN


def test_sanitize_sheet_name_dedupes_when_truncation_causes_collision():
    # Dos nombres que quedan idénticos tras truncar a 31 deben terminar distintos.
    base = "A" * 40
    existing = {base[:MAX_SHEET_NAME_LEN]}
    name = sanitize_sheet_name(base, existing_names=existing)
    assert name not in existing
    assert len(name) <= MAX_SHEET_NAME_LEN


def test_sanitize_sheet_name_falls_back_when_empty_after_sanitizing():
    name = sanitize_sheet_name("::::", existing_names=set())
    assert name  # no queda vacío ni None


# --- write_audit_report: corrige la extensión si no es .xlsx (el formato siempre es xlsx) ---


def test_write_audit_report_corrects_non_xlsx_extension(tmp_path, review_cases, capsys):
    requested_path = tmp_path / "audit_report.csv"
    result = write_audit_report(review_cases, CONFIG, requested_path)

    assert result == tmp_path / "audit_report.xlsx"
    assert result.exists()
    assert not requested_path.exists()
    assert ".xlsx" in capsys.readouterr().out


def test_write_audit_report_corrected_path_is_a_real_xlsx_not_mislabeled(tmp_path, review_cases):
    requested_path = tmp_path / "audit_report.csv"
    result = write_audit_report(review_cases, CONFIG, requested_path)

    with open(result, "rb") as f:
        magic = f.read(4)
    assert magic == b"PK\x03\x04"  # firma ZIP/OOXML real, no texto CSV


# --- write_audit_report: caso sin casos para revisión ---


def test_write_audit_report_returns_none_when_no_review_cases(tmp_path):
    output_path = tmp_path / "auditoria.xlsx"
    result = write_audit_report([], CONFIG, output_path)
    assert result is None
    assert not output_path.exists()


def test_write_audit_report_removes_stale_file_when_no_review_cases(tmp_path):
    # Un archivo de una corrida anterior no debe quedar mintiendo sobre "sin casos".
    output_path = tmp_path / "auditoria.xlsx"
    output_path.write_bytes(b"contenido viejo")
    result = write_audit_report([], CONFIG, output_path)
    assert result is None
    assert not output_path.exists()


# --- write_audit_report: con casos reales (jpeg y pdf) ---


@pytest.fixture
def review_cases():
    jpeg_path = BOLETAS_DIR / "Dinner with BCP.jpeg"
    pdf_path = BOLETAS_DIR / "Taxi to Kyndryl.pdf"
    assert jpeg_path.exists() and pdf_path.exists()

    return [
        (
            jpeg_path,
            _make_result(currency=None, amount=363.0),
            ValidationResult(status="REVIEW", reasons=["no se pudo determinar la moneda"]),
        ),
        (
            pdf_path,
            _make_result(amount=23.90, currency="PEN"),
            ValidationResult(status="REVIEW", reasons=["confianza baja en 'amount' (0.40 < 0.6)"]),
        ),
    ]


def test_write_audit_report_creates_one_sheet_per_case_plus_index(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")
    assert output_path is not None
    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    assert INDEX_SHEET_NAME in wb.sheetnames
    # Índice + una pestaña por caso.
    assert len(wb.sheetnames) == 1 + len(review_cases)


def test_write_audit_report_sheet_contains_reason_and_extracted_values(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")
    wb = openpyxl.load_workbook(output_path)

    jpeg_case_sheet = [s for s in wb.sheetnames if s != INDEX_SHEET_NAME][0]
    ws = wb[jpeg_case_sheet]
    cell_values = [ws.cell(row=r, column=c).value for r in range(1, 6) for c in (1, 2)]

    assert "Dinner with BCP.jpeg" in cell_values
    assert "no se pudo determinar la moneda" in cell_values
    assert 363.0 in cell_values


def test_write_audit_report_index_lists_all_cases_with_reasons(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")
    wb = openpyxl.load_workbook(output_path)
    index_ws = wb[INDEX_SHEET_NAME]

    filenames = [index_ws.cell(row=r, column=1).value for r in range(2, 2 + len(review_cases))]
    reasons = [index_ws.cell(row=r, column=2).value for r in range(2, 2 + len(review_cases))]

    assert "Dinner with BCP.jpeg" in filenames
    assert "Taxi to Kyndryl.pdf" in filenames
    assert any("moneda" in r for r in reasons)
    assert any("amount" in r for r in reasons)


def test_write_audit_report_index_links_to_each_sheet(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")
    wb = openpyxl.load_workbook(output_path)
    index_ws = wb[INDEX_SHEET_NAME]

    hyperlinks = [
        index_ws.cell(row=r, column=1).hyperlink for r in range(2, 2 + len(review_cases))
    ]
    assert all(h is not None for h in hyperlinks)


def test_write_audit_report_embeds_original_images_for_jpeg_and_pdf(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")

    with zipfile.ZipFile(output_path) as zf:
        media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]

    # Una imagen embebida por caso: la del jpeg (tal cual) y la del pdf (rasterizado).
    assert len(media_files) == len(review_cases)


def test_write_audit_report_does_not_use_placeholder_text_when_image_available(tmp_path, review_cases):
    output_path = write_audit_report(review_cases, CONFIG, tmp_path / "auditoria.xlsx")
    wb = openpyxl.load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        if sheet_name == INDEX_SHEET_NAME:
            continue
        ws = wb[sheet_name]
        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert not any(
            v and "No se pudo cargar la imagen" in str(v) for v in all_values
        )
