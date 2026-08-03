import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

import cost
import main
import preprocess
from audit_writer import INDEX_SHEET_NAME
from extract import ExtractionResult
from complementary_info import CurrencyConversion
from excel_writer import ExpenseRow
from main import (
    REVIEW_COMMENTS_SUFFIX,
    FxRequirements,
    _build_expense_row,
    _ensure_utf8_console,
    _format_date_like_excel,
    build_comments,
    build_review_comments,
    compute_missing_requirements,
    detect_fx_requirements,
    determine_fx,
    process_all,
    prompt_fx_for_currency,
    prompt_receipt_selection,
    prompt_report_description,
    prompt_usd_charged,
    sanitize_filename_component,
)
from validate import ValidationResult


def _mock_inputs(monkeypatch, responses):
    """Reemplaza input() por una cola de respuestas fijas, en orden."""
    it = iter(responses)
    monkeypatch.setattr("builtins.input", lambda prompt="": next(it))

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = (
    PROJECT_ROOT / "plantilla" / "Form - Chile Expense Report - Peru Travel Junio 2026.xlsx"
)

BASE_CONFIG = {
    "excel": {
        "template_path": str(TEMPLATE_PATH),
        "sheet_name": "Expense Report",
        "first_data_row": 9,
        "last_data_row": 32,
    },
    "extraction": {"currencies": ["CLP", "USD", "BRL", "ARG", "PEN", "COP", "EUR"]},
    "validation": {
        "min_field_confidence": 0.6,
        "min_overall_confidence": 0.6,
        "amount_consistency_tolerance": 1.0,
    },
    "pricing": {
        "usd_per_million": {"input": 3.00, "output": 15.00, "cache_write": 3.75, "cache_read": 0.30}
    },
}


def test_format_date_like_excel_matches_date_column_format():
    # Mismo formato que la columna Date de la plantilla ([$-409]d\-mmm\-yy;@).
    assert _format_date_like_excel(date(2026, 5, 27)) == "27-May-26"


def test_format_date_like_excel_does_not_zero_pad_day():
    # El código de formato de Excel usa "d" (sin ceros a la izquierda), no "dd".
    assert _format_date_like_excel(date(2026, 6, 7)) == "7-Jun-26"


def test_build_comments_strips_extension():
    result = build_comments("Taxi to Kyndryl.pdf")
    assert result == "Taxi to Kyndryl"


def test_build_comments_strips_only_last_extension_with_multiple_dots():
    result = build_comments("boleta.lima.01.jpg")
    assert result == "boleta.lima.01"


def test_run_prints_execution_summary_with_accumulated_usage(tmp_path, monkeypatch, capsys):
    # No golpea la red: mockea preprocess.preprocess_file y extract_receipt para
    # simular dos boletas procesadas, cada una con su propio usage de la API.
    input_dir = tmp_path / "boletas"
    input_dir.mkdir()
    (input_dir / "a.jpg").write_bytes(b"fake")
    (input_dir / "b.jpg").write_bytes(b"fake")

    usages = iter(
        [
            {"input_tokens": 1000, "output_tokens": 150},
            {"input_tokens": 500, "output_tokens": 50, "cache_read_input_tokens": 200},
        ]
    )

    def fake_preprocess_file(file_path, config):
        return ["fake-page"]

    def fake_extract_receipt(pages, config):
        return ExtractionResult(
            date="2026-06-12",
            currency="PEN",
            amount=100.0,
            expense_type="Taxi",
            confidence={
                "overall": 0.95,
                "amount": 0.95,
                "currency": 0.95,
                "date": 0.95,
                "expense_type": 0.95,
            },
            usage=next(usages),
        )

    monkeypatch.setattr(preprocess, "preprocess_file", fake_preprocess_file)
    monkeypatch.setattr(main, "extract_receipt", fake_extract_receipt)
    # Ambas boletas son PEN (no-CLP, no-USD): no hay boleta en USD así que se
    # pregunta el TC USD->CLP aparte, luego la boleta a usar (índice), el USD
    # cobrado por el banco, y por último la descripción del reporte.
    _mock_inputs(monkeypatch, ["922", "1", "50", "resumen-test"])

    config = {
        "excel": {
            "template_path": str(TEMPLATE_PATH),
            "sheet_name": "Expense Report",
            "first_data_row": 9,
            "last_data_row": 32,
        },
        "extraction": {"currencies": ["CLP", "USD", "BRL", "ARG", "PEN", "COP", "EUR"]},
        "validation": {
            "min_field_confidence": 0.6,
            "min_overall_confidence": 0.6,
            "amount_consistency_tolerance": 1.0,
        },
        "pricing": {
            "usd_per_million": {"input": 3.00, "output": 15.00, "cache_write": 3.75, "cache_read": 0.30}
        },
    }

    main.run(input_dir, config, tmp_path / "out.xlsx", tmp_path / "auditoria.xlsx")

    out = capsys.readouterr().out
    assert "── Resumen de ejecución ──" in out
    assert "Boletas procesadas:    2" in out
    assert "Tokens de entrada:     1,500" in out  # 1000 + 500, acumulado entre las dos boletas
    assert "Tokens de salida:      200" in out  # 150 + 50
    assert "Costo estimado (USD):" in out
    assert "estimado" in out.lower()


# --- _build_expense_row: filas para revisión con monto sí se escriben ---


def _make_ok_result(**overrides) -> ExtractionResult:
    defaults = dict(
        date="2026-06-12",
        currency="PEN",
        amount=100.0,
        expense_type="Taxi",
        confidence={
            "overall": 0.95,
            "amount": 0.95,
            "currency": 0.95,
            "date": 0.95,
            "expense_type": 0.95,
        },
    )
    defaults.update(overrides)
    return ExtractionResult(**defaults)


def test_build_review_comments_appends_suffix_to_filename_stem():
    result = build_review_comments("boleta_taxi_01.jpg")
    assert result == "boleta_taxi_01 (Marcada para Revision)"


def test_build_expense_row_review_with_amount_writes_row_with_review_comments():
    # Boleta marcada para revisión (ej. moneda no determinada) pero con monto sí
    # determinado: la fila se escribe, con Comments = nombre sin extensión + sufijo
    # de revisión (para poder confirmarla borrando solo el sufijo).
    result = _make_ok_result(currency=None)
    validation = ValidationResult(status="REVIEW", reasons=["no se pudo determinar la moneda"])

    row = _build_expense_row(Path("boleta.jpg"), result, validation)

    assert row is not None
    assert row.amount == 100.0
    assert row.currency is None
    assert row.comments == "boleta" + REVIEW_COMMENTS_SUFFIX == "boleta (Marcada para Revision)"


def test_build_expense_row_ok_keeps_filename_without_extension_comments():
    result = _make_ok_result()
    validation = ValidationResult(status="OK", reasons=[])

    row = _build_expense_row(Path("Taxi to Kyndryl.pdf"), result, validation)

    assert row is not None
    assert row.comments == "Taxi to Kyndryl"


def test_build_expense_row_review_without_amount_returns_none():
    # Sin monto determinado: comportamiento actual sin cambios, no se escribe fila.
    result = _make_ok_result(amount=None)
    validation = ValidationResult(status="REVIEW", reasons=["no se pudo determinar el monto"])

    row = _build_expense_row(Path("boleta.jpg"), result, validation)

    assert row is None


def test_run_writes_review_rows_with_amount_and_keeps_review_marking(tmp_path, monkeypatch, capsys):
    # Tres boletas: una OK, una para revisión CON monto (debe escribirse con el
    # comment fijo), una para revisión SIN monto (no debe escribirse, como hoy).
    input_dir = tmp_path / "boletas"
    input_dir.mkdir()
    (input_dir / "ok.jpg").write_bytes(b"fake")
    (input_dir / "review_with_amount.jpg").write_bytes(b"fake")
    (input_dir / "review_without_amount.jpg").write_bytes(b"fake")

    results_by_file = {
        "ok.jpg": _make_ok_result(),
        # Moneda no reconocida -> REVIEW, pero el monto sí se determinó.
        "review_with_amount.jpg": _make_ok_result(currency=None, amount=250.0),
        # Sin monto -> REVIEW y sin dato para escribir la fila.
        "review_without_amount.jpg": _make_ok_result(amount=None),
    }

    # Reemplaza process_file completo (evita depender de preprocess/extract_receipt):
    # dado el archivo, devuelve el ExtractionResult fijado arriba y su validación real.
    def fake_process_file(file_path, config):
        result = results_by_file[file_path.name]
        validation = main.validate.validate_extraction(result, config)
        return result, validation

    monkeypatch.setattr(main, "process_file", fake_process_file)
    # Solo ok.jpg queda con moneda determinada (PEN); review_with_amount.jpg tiene
    # currency=None así que no se pregunta FX para ella. PEN no-USD: TC USD->CLP
    # aparte, índice de boleta (única candidata), USD cobrado, y la descripción.
    _mock_inputs(monkeypatch, ["922", "1", "50", "casos-revision"])

    output_path = tmp_path / "out.xlsx"
    audit_path = tmp_path / "auditoria.xlsx"
    main.run(input_dir, BASE_CONFIG, output_path, audit_path)

    out = capsys.readouterr().out
    # (d) la marca de revisión sigue apareciendo en el resumen impreso.
    assert "Resumen: 1 OK, 2 para revisión, 3 total" in out

    # El nombre de archivo se reemplaza siempre por Expense_Report_<descripción>.xlsx,
    # en el mismo directorio que --output (no en el nombre original "out.xlsx").
    renamed_output_path = tmp_path / "Expense_Report_casos-revision.xlsx"
    assert renamed_output_path.exists()
    assert not output_path.exists()

    wb = openpyxl.load_workbook(renamed_output_path)
    ws = wb["Expense Report"]

    written_comments = [
        ws.cell(row=r, column=9).value for r in range(9, 11) if ws.cell(row=r, column=9).value
    ]
    # Solo 2 filas escritas: la OK y la de revisión CON monto (no la de revisión sin monto).
    assert len(written_comments) == 2
    assert "review_with_amount" + REVIEW_COMMENTS_SUFFIX in written_comments
    assert "ok" in written_comments
    assert not any("review_without_amount" in c for c in written_comments)

    # El .xlsx de auditoría tiene una pestaña por boleta REVIEW (ok.jpg no tiene, ya
    # quedó en el Excel de rendición) más la pestaña índice.
    audit_wb = openpyxl.load_workbook(audit_path)
    assert INDEX_SHEET_NAME in audit_wb.sheetnames
    assert len(audit_wb.sheetnames) == 1 + 2
    index_ws = audit_wb[INDEX_SHEET_NAME]
    filenames_in_index = {index_ws.cell(row=r, column=1).value for r in range(2, 4)}
    assert filenames_in_index == {"review_with_amount.jpg", "review_without_amount.jpg"}


# --- sanitize_filename_component ---


def test_sanitize_filename_component_replaces_forbidden_characters():
    result = sanitize_filename_component('Viaje/Peru: reporte*final?"raro"<>|')
    assert not any(ch in result for ch in '/\\:*?"<>|')


def test_sanitize_filename_component_replaces_internal_spaces_with_underscore():
    assert sanitize_filename_component("Viaje a Lima Junio") == "Viaje_a_Lima_Junio"


def test_sanitize_filename_component_trims_surrounding_whitespace():
    assert sanitize_filename_component("  Viaje Lima  ") == "Viaje_Lima"


def test_sanitize_filename_component_collapses_repeated_underscores():
    # Espacios + separadores prohibidos consecutivos no deben dejar "___".
    assert sanitize_filename_component("Viaje //   Lima") == "Viaje_Lima"


def test_sanitize_filename_component_can_result_in_empty_string():
    assert sanitize_filename_component("   ///:::   ") == ""


# --- prompt_report_description ---


def test_prompt_report_description_returns_sanitized_value(monkeypatch):
    _mock_inputs(monkeypatch, ["Viaje a Lima Junio"])
    assert prompt_report_description() == "Viaje_a_Lima_Junio"


def test_prompt_report_description_reprompts_when_sanitized_result_is_empty(monkeypatch, capsys):
    # "   " y "///" sanean a vacío; recién la tercera respuesta es válida.
    _mock_inputs(monkeypatch, ["   ", "///", "reporte-valido"])
    assert prompt_report_description() == "reporte-valido"
    assert "vacía" in capsys.readouterr().out.lower()


# --- prompt_fx_for_currency ---


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("922", 922.0),
        ("922.50", 922.5),
        ("922,50", 922.5),
    ],
)
def test_prompt_fx_for_currency_accepts_dot_or_comma_decimal_separator(monkeypatch, raw, expected):
    _mock_inputs(monkeypatch, [raw])
    assert prompt_fx_for_currency("USD") == expected


def test_prompt_fx_for_currency_reprompts_on_invalid_input(monkeypatch, capsys):
    # No numérico, cero y negativo son inválidos; recién "922,50" es válido.
    _mock_inputs(monkeypatch, ["abc", "0", "-5", "922,50"])
    assert prompt_fx_for_currency("USD") == 922.5
    assert "inválido" in capsys.readouterr().out.lower()


# --- prompt_receipt_selection ---


def _make_row(source_file, amount, currency="PEN", date_value=None, file_path=None):
    return ExpenseRow(
        date=date_value,
        amount=amount,
        currency=currency,
        expense_type="Taxi",
        source_file=source_file,
        file_path=file_path or Path(source_file),
    )


def test_prompt_receipt_selection_returns_row_at_chosen_index(monkeypatch):
    candidates = [_make_row("a.jpg", 100.0), _make_row("b.jpg", 200.0)]
    _mock_inputs(monkeypatch, ["2"])
    assert prompt_receipt_selection("PEN", candidates) is candidates[1]


def test_prompt_receipt_selection_reprompts_on_out_of_range_or_non_numeric(monkeypatch, capsys):
    candidates = [_make_row("a.jpg", 100.0)]
    _mock_inputs(monkeypatch, ["abc", "0", "2", "1"])
    assert prompt_receipt_selection("PEN", candidates) is candidates[0]
    assert "1" in capsys.readouterr().out


# --- prompt_usd_charged ---


def test_prompt_usd_charged_accepts_positive_number(monkeypatch):
    _mock_inputs(monkeypatch, ["67.41"])
    assert prompt_usd_charged("PEN") == 67.41


def test_prompt_usd_charged_reprompts_on_invalid_input(monkeypatch, capsys):
    _mock_inputs(monkeypatch, ["abc", "-1", "67,41"])
    assert prompt_usd_charged("PEN") == 67.41
    assert "inválido" in capsys.readouterr().out.lower()


# --- determine_fx ---


def test_determine_fx_returns_empty_when_only_clp_present(monkeypatch):
    def _fail_if_called(prompt=""):
        raise AssertionError("no debería pedir nada si la única moneda es CLP")

    monkeypatch.setattr("builtins.input", _fail_if_called)
    rows = [_make_row("a.jpg", 33200.0, currency="CLP")]
    fx_rates, conversions = determine_fx(rows)
    assert fx_rates == {}
    assert conversions == []


def test_determine_fx_ignores_none_currency(monkeypatch):
    def _fail_if_called(prompt=""):
        raise AssertionError("no debería pedir nada para moneda no determinada")

    monkeypatch.setattr("builtins.input", _fail_if_called)
    rows = [_make_row("a.jpg", 100.0, currency=None)]
    fx_rates, conversions = determine_fx(rows)
    assert fx_rates == {}
    assert conversions == []


def test_determine_fx_asks_usd_directly_no_real_fx_calc(monkeypatch):
    # Reporte solo con USD: se pregunta directo, sin pasar por selección de boleta.
    rows = [_make_row("a.jpg", 990.0, currency="USD")]
    _mock_inputs(monkeypatch, ["922"])
    fx_rates, conversions = determine_fx(rows)
    assert fx_rates == {"USD": 922.0}
    assert conversions == []


def test_determine_fx_computes_real_fx_for_non_usd_currency_reusing_usd_rate(monkeypatch):
    # Caso del enunciado: boleta PEN 223.50, USD cobrado 67.41, USD presente con FX 922.
    rows = [
        _make_row("usd.jpg", 990.0, currency="USD"),
        _make_row("pen.jpg", 223.50, currency="PEN", file_path=Path("boletas/pen.jpg")),
    ]
    # FX de USD, luego índice de boleta PEN (única candidata), luego USD cobrado.
    _mock_inputs(monkeypatch, ["922", "1", "67.41"])

    fx_rates, conversions = determine_fx(rows)

    assert fx_rates["USD"] == 922.0
    assert fx_rates["PEN"] == pytest.approx(278.09, abs=0.01)
    assert len(conversions) == 1
    conv = conversions[0]
    assert conv.currency == "PEN"
    assert conv.file_path == Path("boletas/pen.jpg")
    assert conv.dato_a == 223.50
    assert conv.dato_b == 67.41
    assert conv.dato_c == 922.0
    assert conv.fx_step1 == pytest.approx(0.301611, abs=1e-5)
    assert conv.fx_step2 == pytest.approx(278.09, abs=0.01)


def test_determine_fx_asks_usd_to_clp_separately_when_no_usd_boleta(monkeypatch, capsys):
    # Ninguna boleta en USD: el TC USD->CLP se pregunta aparte, una sola vez.
    rows = [_make_row("pen.jpg", 223.50, currency="PEN")]
    _mock_inputs(monkeypatch, ["922", "1", "67.41"])

    fx_rates, conversions = determine_fx(rows)

    assert "USD" not in fx_rates
    assert fx_rates["PEN"] == pytest.approx(278.09, abs=0.01)
    assert conversions[0].dato_c == 922.0
    assert "usd" in capsys.readouterr().out.lower()


def test_determine_fx_reuses_single_usd_to_clp_rate_across_multiple_currencies(monkeypatch):
    # PEN y BRL, sin USD presente: se pregunta el TC USD->CLP UNA sola vez y se
    # reutiliza para ambas conversiones (no se pregunta dos veces).
    rows = [
        _make_row("pen.jpg", 223.50, currency="PEN"),
        _make_row("brl.jpg", 100.0, currency="BRL"),
    ]
    # TC USD->CLP, boleta BRL (índice), USD cobrado BRL, boleta PEN (índice), USD cobrado PEN.
    # BRL va antes que PEN alfabéticamente.
    _mock_inputs(monkeypatch, ["922", "1", "20", "1", "67.41"])

    fx_rates, conversions = determine_fx(rows)

    assert {c.currency for c in conversions} == {"PEN", "BRL"}
    assert all(c.dato_c == 922.0 for c in conversions)


# --- detect_fx_requirements: detección de monedas, extraída de determine_fx para
# que la capa web pueda mostrar los mismos campos dinámicos sin pasar por consola ---


def test_detect_fx_requirements_empty_when_only_clp():
    rows = [_make_row("a.jpg", 33200.0, currency="CLP")]
    req = detect_fx_requirements(rows)
    assert req == FxRequirements(has_usd=False, conversion_currencies=[])


def test_detect_fx_requirements_ignores_none_currency():
    rows = [_make_row("a.jpg", 100.0, currency=None)]
    req = detect_fx_requirements(rows)
    assert req == FxRequirements(has_usd=False, conversion_currencies=[])


def test_detect_fx_requirements_flags_usd_and_sorts_other_currencies():
    rows = [
        _make_row("usd.jpg", 990.0, currency="USD"),
        _make_row("brl.jpg", 100.0, currency="BRL"),
        _make_row("pen.jpg", 223.50, currency="PEN"),
    ]
    req = detect_fx_requirements(rows)
    assert req.has_usd is True
    assert req.conversion_currencies == ["BRL", "PEN"]


# --- compute_missing_requirements: regla pura de habilitación del botón "Crear
# rendición" en la interfaz web ---


def test_compute_missing_requirements_empty_when_everything_present():
    req = FxRequirements(has_usd=False, conversion_currencies=[])
    missing = compute_missing_requirements(
        n_files=1,
        parsed=True,
        description="viaje-lima",
        requirements=req,
        usd_fx=None,
        submitted_conversion_currencies=[],
    )
    assert missing == []


def test_compute_missing_requirements_flags_no_files_not_parsed_and_blank_description():
    missing = compute_missing_requirements(
        n_files=0,
        parsed=False,
        description="   ",
        requirements=None,
        usd_fx=None,
        submitted_conversion_currencies=[],
    )
    assert len(missing) == 3


def test_compute_missing_requirements_requires_usd_fx_when_needed():
    req = FxRequirements(has_usd=True, conversion_currencies=[])
    missing = compute_missing_requirements(
        n_files=1,
        parsed=True,
        description="viaje-lima",
        requirements=req,
        usd_fx=None,
        submitted_conversion_currencies=[],
    )
    assert any("USD" in m for m in missing)


def test_compute_missing_requirements_requires_pending_conversion_currencies():
    req = FxRequirements(has_usd=False, conversion_currencies=["PEN", "BRL"])
    missing = compute_missing_requirements(
        n_files=1,
        parsed=True,
        description="viaje-lima",
        requirements=req,
        usd_fx=922.0,
        submitted_conversion_currencies=["PEN"],
    )
    assert len(missing) == 1
    assert "BRL" in missing[0]
    assert "PEN" not in missing[0]


def test_compute_missing_requirements_ready_once_all_conversions_submitted():
    req = FxRequirements(has_usd=False, conversion_currencies=["PEN", "BRL"])
    missing = compute_missing_requirements(
        n_files=1,
        parsed=True,
        description="viaje-lima",
        requirements=req,
        usd_fx=922.0,
        submitted_conversion_currencies=["PEN", "BRL"],
    )
    assert missing == []


# --- process_all: comparte la extracción de main.run() con la capa web ---


def test_process_all_builds_summary_and_sorts_rows(tmp_path, monkeypatch):
    (tmp_path / "b.jpg").write_bytes(b"fake")
    (tmp_path / "a.jpg").write_bytes(b"fake")

    results_by_file = {
        "a.jpg": _make_ok_result(date="2026-06-10"),
        "b.jpg": _make_ok_result(date="2026-06-05"),
    }

    def fake_process_file(file_path, config):
        result = results_by_file[file_path.name]
        validation = main.validate.validate_extraction(result, BASE_CONFIG)
        return result, validation

    monkeypatch.setattr(main, "process_file", fake_process_file)

    progressed = []
    summary = process_all(
        [tmp_path / "b.jpg", tmp_path / "a.jpg"],
        BASE_CONFIG,
        on_progress=lambda fp: progressed.append(fp.name),
    )

    assert summary.n_total == 2
    assert summary.n_ok == 2
    assert summary.review_cases == []
    # Ordenadas por fecha, no por el orden en que se pasaron los archivos.
    assert [r.source_file for r in summary.rows_to_write] == ["b.jpg", "a.jpg"]
    assert progressed == ["b.jpg", "a.jpg"]


# --- run(): el FX pedido por consola se carga en cada fila según su moneda ---


def test_run_loads_prompted_fx_per_currency_and_keeps_clp_at_one(tmp_path, monkeypatch):
    input_dir = tmp_path / "boletas"
    input_dir.mkdir()
    (input_dir / "clp.jpg").write_bytes(b"fake")
    (input_dir / "usd.jpg").write_bytes(b"fake")

    results_by_file = {
        "clp.jpg": _make_ok_result(currency="CLP", amount=33200.0),
        "usd.jpg": _make_ok_result(currency="USD", amount=990.0),
    }

    def fake_process_file(file_path, config):
        result = results_by_file[file_path.name]
        validation = main.validate.validate_extraction(result, config)
        return result, validation

    monkeypatch.setattr(main, "process_file", fake_process_file)
    # Única moneda no-CLP presente: USD. FX ingresado = 922,50 (con coma decimal).
    _mock_inputs(monkeypatch, ["922,50", "viaje-mixto"])

    output_path = tmp_path / "out.xlsx"
    main.run(input_dir, BASE_CONFIG, output_path, tmp_path / "auditoria.xlsx")

    wb = openpyxl.load_workbook(tmp_path / "Expense_Report_viaje-mixto.xlsx")
    ws = wb["Expense Report"]

    rows_by_currency = {
        ws.cell(row=r, column=5).value: r
        for r in (9, 10)
        if ws.cell(row=r, column=5).value
    }
    clp_row = rows_by_currency["CLP"]
    usd_row = rows_by_currency["USD"]

    assert ws.cell(row=clp_row, column=6).value == 1  # FX de CLP siempre 1, no se pregunta
    assert ws.cell(row=usd_row, column=6).value == 922.5  # FX ingresado por el usuario
    # Amount in CLP se mantiene como fórmula FX×Amount, no un número fijo.
    assert ws.cell(row=usd_row, column=7).value == f"=F{usd_row}*D{usd_row}"


# --- _ensure_utf8_console: compatibilidad con consola Windows (code page legacy) ---


def test_ensure_utf8_console_reconfigures_stdout_and_stderr(monkeypatch):
    calls = []

    class FakeStream:
        def reconfigure(self, encoding=None):
            calls.append(encoding)

    monkeypatch.setattr(main.sys, "stdout", FakeStream())
    monkeypatch.setattr(main.sys, "stderr", FakeStream())

    _ensure_utf8_console()

    assert calls == ["utf-8", "utf-8"]


def test_ensure_utf8_console_tolerates_streams_without_reconfigure(monkeypatch):
    # Streams redirigidos/capturados que no exponen reconfigure() no deben hacer
    # fallar el arranque del programa.
    class FakeStreamNoReconfigure:
        pass

    monkeypatch.setattr(main.sys, "stdout", FakeStreamNoReconfigure())
    monkeypatch.setattr(main.sys, "stderr", FakeStreamNoReconfigure())

    _ensure_utf8_console()  # no debe lanzar


def test_ensure_utf8_console_fixes_unicode_encode_error_on_legacy_codepage(monkeypatch):
    # cp1252 (code page legacy común en Windows) no puede codificar "─" (guion
    # largo del resumen de ejecución) ni varias tildes — reproduce el error real
    # antes de la corrección, y confirma que reconfigurar a utf-8 lo resuelve.
    summary = cost.format_summary(1, cost.TokenUsage(input_tokens=10, output_tokens=2), {})

    buffer_before = io.BytesIO()
    stream_before = io.TextIOWrapper(buffer_before, encoding="cp1252")
    with pytest.raises(UnicodeEncodeError):
        stream_before.write(summary)
        stream_before.flush()

    buffer_after = io.BytesIO()
    stream_after = io.TextIOWrapper(buffer_after, encoding="cp1252")
    monkeypatch.setattr(main.sys, "stdout", stream_after)
    monkeypatch.setattr(main.sys, "stderr", stream_after)

    _ensure_utf8_console()

    main.sys.stdout.write(summary)
    main.sys.stdout.flush()
    assert "Resumen de ejecución" in buffer_after.getvalue().decode("utf-8")
