"""Tests de la capa web (src/web/routes.py) vía FastAPI TestClient.

No pegan a la API de Claude: igual que en test_main.py, se mockea
`main.process_file` para simular la extracción. El objetivo es confirmar que la
capa HTTP es una fachada correcta sobre el mismo pipeline (main/excel_writer/
audit_writer/complementary_info) — no reprobar esa lógica de negocio, que ya
tiene su propia cobertura en los demás archivos de test.
"""

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

import main
from extract import ExtractionResult
from web import routes


@pytest.fixture(autouse=True)
def _isolated_secrets(tmp_path, monkeypatch):
    # Evita depender del secrets.yaml / ANTHROPIC_API_KEY real de la máquina.
    monkeypatch.setattr(routes, "SECRETS_PATH", tmp_path / "secrets.yaml")
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)


@pytest.fixture
def client():
    return TestClient(routes.app)


def _make_ok_result(**overrides) -> ExtractionResult:
    defaults = dict(
        date="2026-06-12",
        currency="PEN",
        amount=100.0,
        expense_type="Taxi",
        confidence={
            "overall": 0.95,
            "amount": 0.95,
            "currency": 0.95,
            "date": 0.95,
            "expense_type": 0.95,
        },
    )
    defaults.update(overrides)
    return ExtractionResult(**defaults)


def _create_job(client) -> str:
    return client.post("/api/jobs").json()["job_id"]


def _upload(client, job_id, filenames):
    files = [("files", (name, io.BytesIO(b"fake-bytes"), "image/jpeg")) for name in filenames]
    return client.post(f"/api/jobs/{job_id}/receipts", files=files)


def _mock_results(monkeypatch, results_by_file, config=None):
    cfg = config or main.load_config(routes.CONFIG_PATH)

    def fake_process_file(file_path, _config):
        result = results_by_file[file_path.name]
        validation = main.validate.validate_extraction(result, cfg)
        return result, validation

    monkeypatch.setattr(main, "process_file", fake_process_file)


# --- página y config ---------------------------------------------------------


def test_index_serves_html(client):
    r = client.get("/")
    assert r.status_code == 200
    assert "text/html" in r.headers["content-type"]
    assert "HITACHI" in r.text


def test_config_status_false_when_no_key(client):
    r = client.get("/api/config-status")
    assert r.json() == {"has_api_key": False}


def test_config_status_true_after_env_var_set(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    r = client.get("/api/config-status")
    assert r.json() == {"has_api_key": True}


def test_set_api_key_persists_and_updates_status(client, tmp_path):
    r = client.post("/api/config/api-key", json={"api_key": "sk-ant-test"})
    assert r.status_code == 200
    assert r.json() == {"has_api_key": True}

    r = client.get("/api/config-status")
    assert r.json() == {"has_api_key": True}

    saved = (routes.SECRETS_PATH).read_text(encoding="utf-8")
    assert "sk-ant-test" in saved


def test_set_api_key_rejects_blank(client):
    r = client.post("/api/config/api-key", json={"api_key": "   "})
    assert r.status_code == 400


# --- upload / remove receipts -------------------------------------------------


def test_upload_accepts_supported_and_rejects_others(client):
    job_id = _create_job(client)
    r = _upload(client, job_id, ["a.jpg", "b.txt", "c.pdf"])
    data = r.json()
    assert sorted(data["files"]) == ["a.jpg", "c.pdf"]
    assert data["rejected"] == ["b.txt"]


def test_remove_receipt(client):
    job_id = _create_job(client)
    _upload(client, job_id, ["a.jpg", "b.jpg"])
    r = client.delete(f"/api/jobs/{job_id}/receipts/a.jpg")
    assert r.json()["files"] == ["b.jpg"]


def test_upload_unknown_job_returns_404(client):
    r = _upload(client, "does-not-exist", ["a.jpg"])
    assert r.status_code == 404


# --- parse ---------------------------------------------------------------------


def test_parse_requires_files(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    r = client.post(f"/api/jobs/{job_id}/parse")
    assert r.status_code == 400


def test_parse_requires_api_key(client):
    job_id = _create_job(client)
    _upload(client, job_id, ["a.jpg"])
    r = client.post(f"/api/jobs/{job_id}/parse")
    assert r.status_code == 400
    assert "API key" in r.json()["detail"]


def test_parse_detects_currencies_and_candidates(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["clp.jpg", "pen.jpg"])

    _mock_results(
        monkeypatch,
        {
            "clp.jpg": _make_ok_result(currency="CLP", amount=33200.0),
            "pen.jpg": _make_ok_result(currency="PEN", amount=223.50),
        },
    )

    r = client.post(f"/api/jobs/{job_id}/parse")
    assert r.status_code == 200
    data = r.json()

    assert data["n_total"] == 2
    assert data["n_ok"] == 2
    assert data["n_review"] == 0
    assert data["needs_usd_fx"] is True  # PEN necesita FX real -> pide Dato C (USD->CLP)
    assert data["conversion_currencies"] == ["PEN"]
    assert data["candidates_by_currency"]["PEN"] == [
        {"filename": "pen.jpg", "amount": 223.50, "date": "2026-06-12"}
    ]
    assert data["usage"]["total_tokens"] == 0  # sin usage real (mock no la puebla)


def test_parse_with_only_clp_needs_no_fx(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["clp.jpg"])
    _mock_results(monkeypatch, {"clp.jpg": _make_ok_result(currency="CLP", amount=33200.0)})

    r = client.post(f"/api/jobs/{job_id}/parse")
    data = r.json()
    assert data["needs_usd_fx"] is False
    assert data["conversion_currencies"] == []


# --- validate (regla de habilitación del botón) --------------------------------


def test_validate_not_ready_before_parse(client):
    job_id = _create_job(client)
    r = client.post(f"/api/jobs/{job_id}/validate", json={"description": "viaje"})
    data = r.json()
    assert data["ready"] is False
    assert data["missing"]


def test_validate_ready_once_description_and_conversions_present(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["pen.jpg"])
    _mock_results(monkeypatch, {"pen.jpg": _make_ok_result(currency="PEN", amount=223.50)})
    client.post(f"/api/jobs/{job_id}/parse")

    r = client.post(
        f"/api/jobs/{job_id}/validate",
        json={"description": "viaje", "usd_fx": None},
    )
    assert r.json()["ready"] is False  # falta usd_fx y la selección de boleta PEN

    r = client.post(
        f"/api/jobs/{job_id}/validate",
        json={
            "description": "viaje",
            "usd_fx": 922,
            "conversions": [{"currency": "PEN", "filename": "pen.jpg", "usd_charged": 67.41}],
        },
    )
    assert r.json() == {"ready": True, "missing": []}


# --- generate --------------------------------------------------------------------


def test_generate_rejects_when_requirements_missing(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["clp.jpg"])
    _mock_results(monkeypatch, {"clp.jpg": _make_ok_result(currency="CLP", amount=33200.0)})
    client.post(f"/api/jobs/{job_id}/parse")

    r = client.post(f"/api/jobs/{job_id}/generate", json={"description": ""})
    assert r.status_code == 400


def test_generate_clp_only_removes_complementary_info_tab_and_offers_downloads(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["clp.jpg"])
    _mock_results(monkeypatch, {"clp.jpg": _make_ok_result(currency="CLP", amount=33200.0)})
    client.post(f"/api/jobs/{job_id}/parse")

    r = client.post(f"/api/jobs/{job_id}/generate", json={"description": "viaje CLP"})
    assert r.status_code == 200
    data = r.json()
    assert data["audit_url"] is None  # sin casos para revisión
    assert data["summary"] == {"n_ok": 1, "n_review": 0, "n_total": 1}

    dl = client.get(data["report_url"])
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    assert "Complementary info" not in wb.sheetnames
    assert not any(name.startswith("Complementary info") for name in wb.sheetnames)

    # No se generó auditoría: el endpoint de descarga debe devolver 404.
    assert client.get(f"/api/jobs/{job_id}/download/audit").status_code == 404


def test_generate_with_conversion_currency_creates_complementary_info_tab(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["pen.jpg"])
    _mock_results(monkeypatch, {"pen.jpg": _make_ok_result(currency="PEN", amount=223.50)})
    client.post(f"/api/jobs/{job_id}/parse")

    r = client.post(
        f"/api/jobs/{job_id}/generate",
        json={
            "description": "viaje lima",
            "usd_fx": 922,
            "conversions": [{"currency": "PEN", "filename": "pen.jpg", "usd_charged": 67.41}],
        },
    )
    assert r.status_code == 200
    data = r.json()

    dl = client.get(data["report_url"])
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    assert "Complementary info - PEN" in wb.sheetnames

    ws = wb["Expense Report"]
    # FX real calculado (223.50 PEN / 67.41 USD * 922) ~ 278.09, cargado en la fila.
    assert ws.cell(row=9, column=6).value == pytest.approx(278.09, abs=0.01)


def test_generate_writes_review_case_and_audit_report_available(client, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "env-token")
    job_id = _create_job(client)
    _upload(client, job_id, ["review.jpg"])
    _mock_results(monkeypatch, {"review.jpg": _make_ok_result(currency=None, amount=250.0)})
    parse_res = client.post(f"/api/jobs/{job_id}/parse").json()
    assert parse_res["n_review"] == 1

    r = client.post(f"/api/jobs/{job_id}/generate", json={"description": "revision"})
    data = r.json()
    assert data["audit_url"] is not None

    dl = client.get(data["audit_url"])
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    assert "Índice" in wb.sheetnames
