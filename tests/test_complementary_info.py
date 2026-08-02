import zipfile
from datetime import date
from pathlib import Path

import openpyxl
import pytest

import excel_writer
from complementary_info import (
    CELL_DATO_A,
    CELL_DATO_B,
    CELL_DATO_C,
    CELL_STEP1_COPY,
    CurrencyConversion,
    compute_real_fx,
)
from excel_writer import ExpenseRow

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = (
    PROJECT_ROOT / "plantilla" / "Form - Chile Expense Report - Peru Travel Junio 2026.xlsx"
)
BOLETAS_DIR = PROJECT_ROOT / "boletas"
RECEIPT_PATH = BOLETAS_DIR / "Coffee with TDSynnex.jpeg"

CONFIG = {
    "excel": {
        "template_path": str(TEMPLATE_PATH),
        "sheet_name": "Expense Report",
        "first_data_row": 9,
        "last_data_row": 32,
    }
}

# --- compute_real_fx: caso real del enunciado (boleta PEN 223.50 / USD 67.41 / TC 922) ---


def test_compute_real_fx_matches_reference_example():
    fx_step1, fx_step2 = compute_real_fx(223.50, 67.41, 922)
    assert fx_step1 == pytest.approx(0.301611, abs=1e-5)
    assert fx_step2 == pytest.approx(278.09, abs=0.01)


def test_currency_conversion_properties_match_compute_real_fx():
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    fx_step1, fx_step2 = compute_real_fx(223.50, 67.41, 922)
    assert conversion.fx_step1 == fx_step1
    assert conversion.fx_step2 == fx_step2


def test_currency_conversion_sheet_title():
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=1, dato_b=1, dato_c=1
    )
    assert conversion.sheet_title() == "Complementary info - PEN"


# --- write_expense_report + conversions: integración contra la plantilla real ---


@pytest.fixture
def receipt_row():
    assert RECEIPT_PATH.exists()
    return ExpenseRow(
        date=date(2026, 6, 12),
        amount=223.50,
        currency="PEN",
        expense_type="Travel - Meals",
        source_file=RECEIPT_PATH.name,
        fx=278.09,
        file_path=RECEIPT_PATH,
    )


def test_write_expense_report_deletes_complementary_info_when_no_conversions(tmp_path, receipt_row):
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report([receipt_row], CONFIG, output_path, conversions=[])

    wb = openpyxl.load_workbook(output_path)
    assert "Complementary info" not in wb.sheetnames
    assert not any(name.startswith("Complementary info") for name in wb.sheetnames)


def test_write_expense_report_creates_one_tab_per_conversion(tmp_path, receipt_row):
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [receipt_row], CONFIG, output_path, conversions=[conversion]
    )

    wb = openpyxl.load_workbook(output_path)
    assert "Complementary info" not in wb.sheetnames
    assert "Complementary info - PEN" in wb.sheetnames

    ws = wb["Complementary info - PEN"]
    assert ws[CELL_DATO_A].value == 223.50
    assert ws[CELL_DATO_B].value == 67.41
    assert ws[CELL_DATO_C].value == 922.0
    assert ws[CELL_STEP1_COPY].value == "=R24"
    # Las fórmulas de la plantilla se preservan tal cual (no las tocamos).
    assert ws["R24"].value == "=Q24/P24"
    assert ws["R30"].value == "=Q30*P31"


def test_write_expense_report_fx_result_matches_expense_report_decimal_format(
    tmp_path, receipt_row
):
    # La plantilla trae R30 (resultado final del FX real) con formato "0" (sin
    # decimales), inconsistente con la columna FX de "Expense Report" (2 decimales) —
    # donde termina usándose ese mismo número. Deben verse con la misma cantidad de
    # decimales en las dos hojas.
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [receipt_row], CONFIG, output_path, conversions=[conversion]
    )

    wb = openpyxl.load_workbook(output_path)
    fx_col_format = wb["Expense Report"]["F9"].number_format
    r30_format = wb["Complementary info - PEN"]["R30"].number_format
    assert r30_format == fx_col_format
    assert r30_format != "0"


def test_write_expense_report_places_complementary_tab_right_after_expense_report(
    tmp_path, receipt_row
):
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [receipt_row], CONFIG, output_path, conversions=[conversion]
    )

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames[0] == "Expense Report"
    assert wb.sheetnames[1] == "Complementary info - PEN"


def test_write_expense_report_embeds_selected_receipt_image_not_placeholder(tmp_path, receipt_row):
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [receipt_row], CONFIG, output_path, conversions=[conversion]
    )

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Complementary info - PEN"]
    # 3 imágenes: la boleta seleccionada (reemplazada) + las 2 del banco (sin tocar).
    assert len(ws._images) == 3


def test_write_expense_report_reinjects_arrows_into_complementary_tab_drawing(tmp_path, receipt_row):
    conversion = CurrencyConversion(
        currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922
    )
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [receipt_row], CONFIG, output_path, conversions=[conversion]
    )

    with zipfile.ZipFile(output_path) as zf:
        drawing_files = [n for n in zf.namelist() if n.startswith("xl/drawings/drawing")]
        found_arrows = any("cxnSp" in zf.read(n).decode("utf-8") for n in drawing_files)
    assert found_arrows


def test_write_expense_report_multiple_currencies_create_separate_tabs(tmp_path):
    pen_receipt = ExpenseRow(
        date=date(2026, 6, 12),
        amount=223.50,
        currency="PEN",
        expense_type="Travel - Meals",
        source_file="pen.jpg",
        fx=278.09,
        file_path=RECEIPT_PATH,
    )
    brl_receipt = ExpenseRow(
        date=date(2026, 6, 13),
        amount=100.0,
        currency="BRL",
        expense_type="Travel - Meals",
        source_file="brl.jpg",
        fx=180.0,
        file_path=RECEIPT_PATH,
    )
    conversions = [
        CurrencyConversion(currency="PEN", file_path=RECEIPT_PATH, dato_a=223.50, dato_b=67.41, dato_c=922),
        CurrencyConversion(currency="BRL", file_path=RECEIPT_PATH, dato_a=100.0, dato_b=20.0, dato_c=922),
    ]
    output_path = tmp_path / "out.xlsx"
    excel_writer.write_expense_report(
        [pen_receipt, brl_receipt], CONFIG, output_path, conversions=conversions
    )

    wb = openpyxl.load_workbook(output_path)
    assert "Complementary info - PEN" in wb.sheetnames
    assert "Complementary info - BRL" in wb.sheetnames
    assert "Complementary info" not in wb.sheetnames
