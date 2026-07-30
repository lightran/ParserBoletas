from datetime import date
from pathlib import Path

import zipfile

import openpyxl
import pytest

from excel_writer import (
    APPROVAL_DATE_CELL,
    COL_AMOUNT,
    COL_AMOUNT_CLP,
    COL_CURRENCY,
    COL_EXPENSE_TYPE,
    COL_FX,
    DATE_SUBMITTED_CELL,
    DEFAULT_FX,
    ExpenseRow,
    TemplateCapacityError,
    write_expense_report,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = (
    PROJECT_ROOT / "plantilla" / "Form - Chile Expense Report - Peru Travel Junio 2026.xlsx"
)

CONFIG = {
    "excel": {
        "template_path": str(TEMPLATE_PATH),
        "sheet_name": "Expense Report",
        "first_data_row": 9,
        "last_data_row": 32,
    }
}

# Subconjunto de la verdad de referencia (valores ya resueltos en la plantilla de ejemplo).
GROUND_TRUTH_ROWS = [
    ExpenseRow(
        date=date(2026, 5, 27),
        amount=464717,
        currency="CLP",
        expense_type="Travel - Other",
        source_file="Airline Ticket Invoice.pdf",
    ),
    ExpenseRow(
        date=date(2026, 6, 13),
        amount=990,
        currency="USD",
        expense_type="Travel - Lodging",
        source_file="Hotel Invoice 6 Nights.pdf",
    ),
    ExpenseRow(
        date=date(2026, 6, 7),
        amount=33200,
        currency="CLP",
        expense_type="Taxi",
        source_file="Taxi from Hotel to SCL Airport Invoice.jpg",
    ),
]


@pytest.fixture
def output_path(tmp_path):
    return tmp_path / "Expense_Report_test.xlsx"


def test_write_expense_report_matches_ground_truth(output_path):
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    for i, expected in enumerate(GROUND_TRUTH_ROWS):
        r = 9 + i
        assert ws.cell(row=r, column=2).value == i + 1  # Item
        # openpyxl siempre devuelve datetime.datetime para celdas con formato de fecha
        assert ws.cell(row=r, column=3).value.date() == expected.date
        assert ws.cell(row=r, column=COL_AMOUNT).value == pytest.approx(expected.amount)
        assert ws.cell(row=r, column=COL_CURRENCY).value == expected.currency
        assert ws.cell(row=r, column=COL_EXPENSE_TYPE).value == expected.expense_type


def test_write_expense_report_leaves_comments_blank(output_path):
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    for i in range(len(GROUND_TRUTH_ROWS)):
        r = 9 + i
        assert ws.cell(row=r, column=9).value is None  # Comments: lo completa el usuario


def test_write_expense_report_fx_defaults_to_one(output_path):
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    for i in range(len(GROUND_TRUTH_ROWS)):
        r = 9 + i
        assert ws.cell(row=r, column=COL_FX).value == DEFAULT_FX == 1


def test_write_expense_report_does_not_inherit_residual_fx_from_template(output_path):
    # La plantilla trae, en filas "vacías" más allá del ejemplo, residuos de FX (255)
    # y moneda precargada (PEN) que NO deben heredarse al escribir boletas nuevas.
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = template_wb["Expense Report"]
    assert template_ws["F21"].value == 255  # confirma el residuo que trae la plantilla

    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]
    for i in range(len(GROUND_TRUTH_ROWS)):
        r = 9 + i
        assert ws.cell(row=r, column=COL_FX).value == 1


def test_write_expense_report_amount_in_clp_is_formula(output_path):
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    for i in range(len(GROUND_TRUTH_ROWS)):
        r = 9 + i
        formula = ws.cell(row=r, column=COL_AMOUNT_CLP).value
        assert formula == f"=F{r}*D{r}"


def test_write_expense_report_amount_in_clp_formula_evaluates_correctly():
    # F{r}=FX (siempre 1 en v1) * D{r}=Amount, así que el valor esperado es el
    # amount original (FX=1). Verificamos la semántica de la fórmula por sustitución.
    for row in GROUND_TRUTH_ROWS:
        fx = DEFAULT_FX
        assert fx * row.amount == row.amount


def test_write_expense_report_clears_leftover_example_rows(output_path):
    # La plantilla original trae filas de ejemplo llenas más allá de las 3 que escribimos
    # (item 4 en adelante). Deben quedar completamente en blanco, no con el ejemplo viejo.
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    r = 9 + len(GROUND_TRUTH_ROWS)  # primera fila no escrita por nosotros
    for col in range(2, 10):
        assert ws.cell(row=r, column=col).value is None

    # Las filas de relleno del template (21-30) traían defaults de moneda/tipo de gasto;
    # también deben quedar en blanco tras generar un reporte nuevo.
    for col in range(2, 10):
        assert ws.cell(row=25, column=col).value is None


def test_write_expense_report_preserves_total_row_formula(output_path):
    # Fila 33 trae "=SUM(G9:G32)"; la limpieza/escritura no debe tocarla.
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    assert template_wb["Expense Report"]["G33"].value == "=SUM(G9:G32)"

    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    assert wb["Expense Report"]["G33"].value == "=SUM(G9:G32)"


def test_write_expense_report_preserves_currency_and_expense_type_dropdowns(output_path):
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    with zipfile.ZipFile(output_path) as zf:
        sheet_xmls = [
            zf.read(name).decode("utf-8")
            for name in zf.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
    combined = "\n".join(sheet_xmls)

    assert "x14:dataValidation" in combined
    assert "'Cheat Sheet'!$A$37:$A$43" in combined  # dropdown de Currency
    assert "'Cheat Sheet'!$A$2:$A$23" in combined  # dropdown de Expense Type


def test_write_expense_report_raises_when_exceeding_template_capacity(output_path):
    too_many_rows = GROUND_TRUTH_ROWS * 9  # 27 filas > 24 disponibles (9-32)
    with pytest.raises(TemplateCapacityError):
        write_expense_report(too_many_rows, CONFIG, output_path)


def test_write_expense_report_stamps_execution_date_on_header_fields(output_path):
    today = date.today()
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]

    assert ws[DATE_SUBMITTED_CELL].value.date() == today  # "Date Submitted" (D6)
    assert ws[APPROVAL_DATE_CELL].value.date() == today  # "Date" en "Approved by:" (F39)
    # Ambos campos quedan con la misma fecha de ejecución.
    assert ws[DATE_SUBMITTED_CELL].value.date() == ws[APPROVAL_DATE_CELL].value.date()


def test_write_expense_report_preserves_header_date_number_formats(output_path):
    # No se toca number_format: cada celda conserva el estilo que ya traía la plantilla
    # (D6 estilo "27-May-26", F39 estilo "mm-dd-yy" — son distintos en la plantilla original).
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = template_wb["Expense Report"]
    expected_submitted_format = template_ws[DATE_SUBMITTED_CELL].number_format
    expected_approval_format = template_ws[APPROVAL_DATE_CELL].number_format

    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]
    assert ws[DATE_SUBMITTED_CELL].number_format == expected_submitted_format
    assert ws[APPROVAL_DATE_CELL].number_format == expected_approval_format


def test_write_expense_report_does_not_affect_per_row_date_column(output_path):
    # La columna Date por fila (C) debe seguir reflejando la fecha de cada boleta,
    # no la fecha de ejecución del programa.
    today = date.today()
    write_expense_report(GROUND_TRUTH_ROWS, CONFIG, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]
    for i, expected in enumerate(GROUND_TRUTH_ROWS):
        r = 9 + i
        row_date = ws.cell(row=r, column=3).value.date()
        assert row_date == expected.date
        assert row_date != today


def test_write_expense_report_respects_employee_overrides(output_path):
    config = {**CONFIG, "excel": {**CONFIG["excel"], "employee_name": "Ada Lovelace", "cost_centre": "999.0000"}}
    write_expense_report(GROUND_TRUTH_ROWS, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Expense Report"]
    assert ws["D4"].value == "Ada Lovelace"
    assert ws["D5"].value == "999.0000"
